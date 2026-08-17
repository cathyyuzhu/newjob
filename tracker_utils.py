"""
JD匹配追踪表 - 可复用工具函数

用途：每次分析完一个新JD后，调用 add_entry() 把结果作为新的一行
插入到追踪表的最上方（row 2，紧跟表头下面），已有记录自动下移。

固定保存路径：此电脑/下载/JD匹配追踪表.xlsx
（Windows下的实际路径形如 C:\\Users\\<用户名>\\Downloads\\JD匹配追踪表.xlsx，
Mac下形如 /Users/<用户名>/Downloads/JD匹配追踪表.xlsx，
调用前请确认当前系统下"下载"文件夹的真实路径）

用法示例：
    from tracker_utils import add_entry

    add_entry(
        path="/path/to/JD匹配追踪表.xlsx",
        job_title="Product Lead, AI Neobank App",
        job_url="https://www.linkedin.com/jobs/view/4454342546/",  # 优先用投递/申请链接(apply link)；如果只有职位展示页链接也可以，但场景三(Indeed)必须用get_job_details返回的apply链接
        company="BJAK",
        overall_match=0.65,   # 0~1之间的小数
        job_content_bullets=[
            "统筹AI Neobank App核心产品方向",
            "覆盖保险、支付、储蓄、投资、旅行等多个金融服务模块",
        ],
        requirement_items=[
            ("资深数字产品负责人经验，能带领多产品方向/squad", False),  # False=已匹配
            ("Fintech/保险/支付/消费应用行业经验", True),               # True=未达标->红色
        ],
        skill_matched_bullets=["领导DemandTec四条产品线，多workstream统筹经验"],
        skill_gap_bullets=["缺少金融类消费者产品简化的直接案例"],
        experience_years="12+年，Director级别，满足资深要求",
        industry_bullets=["无fintech/保险/支付直接经验", "核心背景为零售定价SaaS（最大差距点）"],
        salary="JD未公开，需询问",
        team_bullets=["现任Director，管理多条产品线", "JD未说明具体团队规模"],
        location="远程（需常驻中国）",
        status="低于70%阈值，待用户决定是否生成尝试性简历",
        apply_date=None,  # 默认today
        resume_optimization_bullets=None,  # 情况B/未生成简历时不传，留空即可
        resume_path=None,
        cover_letter=None,
    )

    # 情况A（匹配度达标且生成了定制简历+cover letter）示例：
    add_entry(
        ...,
        resume_optimization_bullets=[
            "Professional Summary第一条改写为突出AI Agent 0-to-1经验",
            "DemandTec经历拆分强调跨职能leadership与交付质量",
        ],
        resume_path="C:\\Users\\Cathy\\Downloads\\Cathy_Yang_Resume_EN_BJAK.docx",
        cover_letter="Dear Hiring Team,\n\nI'm writing to express my interest in...\n\nBest regards,\nCathy Yang",
    )
"""

import os
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

FONT_NAME = "Arial"

HEADERS = [
    "职位名称", "公司", "总体匹配度", "投递日期",
    "职位内容", "任职要求",
    "技能匹配度-匹配的", "技能匹配度-未达标的",
    "相关经验年限", "行业背景", "薪资范围",
    "团队规模/汇报线", "地理位置/远程要求", "状态/下一步",
    "简历优化内容", "简历存储路径", "Cover Letter",
    "公司简介",
]

COL_WIDTHS = {
    "A": 26, "B": 12, "C": 12, "D": 12, "E": 34, "F": 42,
    "G": 30, "H": 28, "I": 18, "J": 26, "K": 16, "L": 26, "M": 18, "N": 26,
    "O": 34, "P": 30, "Q": 50, "R": 40,
}

# 重要：颜色必须用8位ARGB且alpha=FF（不透明）。
# 只传6位RGB（如"000000"）会被openpyxl解读为alpha=00，也就是完全透明，
# 文字会实际上不可见——这是个真实踩过的坑，务必保留FF前缀。
BLACK_RUN = InlineFont(rFont=FONT_NAME, sz=10, color="FF000000")
RED_RUN = InlineFont(rFont=FONT_NAME, sz=10, color="FFFF0000", b=True)

THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)


def _bullets_plain(lines):
    """普通要点列表（黑色，无高亮需求时用）-> 多行字符串"""
    if isinstance(lines, str):
        return lines
    return "\n".join(f"• {l}" for l in lines)


def _bullets_rich(items):
    """items: list of (text, is_gap:bool) -> CellRichText
    每个要点独立一行；未达标(is_gap=True)的要点标红加粗。
    换行符直接拼在每个TextBlock文本末尾（而非单独的换行run）——
    如果用独立的"\\n"-only run，Excel按XML默认空白处理规则会把它丢弃，
    导致所有要点被合并显示成一行。这是实测踩过的坑，务必保持这个写法。
    """
    blocks = []
    n = len(items)
    for i, (text, is_gap) in enumerate(items):
        run = RED_RUN if is_gap else BLACK_RUN
        suffix = "\n" if i < n - 1 else ""
        blocks.append(TextBlock(run, f"• {text}{suffix}"))
    return CellRichText(*blocks)


def _new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "JD匹配追踪表"
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    return wb


def _migrate_headers(ws):
    """给已存在的旧追踪表补上后续新增的表头列（只在行末追加，不改动已有列的位置/内容），
    这样老文件不需要手动删掉重建就能兼容新加的字段（比如后来加的"公司简介"列）。"""
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx)
        if c.value == h:
            continue
        if c.value:  # 已有其它内容（不该发生，HEADERS只在末尾追加），跳过避免误覆盖
            continue
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        col_letter = c.column_letter
        if col_letter in COL_WIDTHS:
            ws.column_dimensions[col_letter].width = COL_WIDTHS[col_letter]


def _ensure_workbook(path):
    if os.path.exists(path):
        wb = load_workbook(path, rich_text=True)
        ws = wb["JD匹配追踪表"] if "JD匹配追踪表" in wb.sheetnames else wb.active
        _migrate_headers(ws)
        return wb
    return _new_workbook()


def add_entry(
    path,
    job_title,
    job_url,
    company,
    overall_match,          # float 0~1
    job_content_bullets,    # list[str]
    requirement_items,      # list[(str, bool)]  bool=True表示未达标->红色
    skill_matched_bullets,  # list[str]
    skill_gap_bullets,      # list[str]
    experience_years,       # str（通常单条，不用bullet）
    industry_bullets,       # list[str] 或 str
    salary,                 # str
    team_bullets,           # list[str] 或 str
    location,               # str，如"远程（需常驻中国）"或具体城市名
    status,                 # str
    apply_date=None,
    resume_optimization_bullets=None,  # list[str]，简历改了哪些地方；情况B/未生成简历时传None
    resume_path=None,                  # str，定制简历的完整保存路径；未生成时传None
    cover_letter=None,                 # str，cover letter全文（含换行），未生成时传None
    company_overview=None,             # str，公司简介（LLM基于自身知识生成的简要介绍）；未知时传None
):
    wb = _ensure_workbook(path)
    ws = wb["JD匹配追踪表"] if "JD匹配追踪表" in wb.sheetnames else wb.active

    # 重新分析同一个职位（公司+职位名称一致）时，先删掉旧行再插入新行，避免表里出现
    # 同一职位的重复记录——重复行不仅让表格变乱，list_entries() 读回来构建
    # company+title -> 记录 的索引时还会被旧行覆盖，导致网页弹窗反而显示回退分析前的
    # 旧数据（比如翻译成中文之前的英文内容）。
    company_col = HEADERS.index("公司") + 1
    title_col = HEADERS.index("职位名称") + 1
    for row in reversed(list(ws.iter_rows(min_row=2, max_row=ws.max_row))):
        if row[company_col - 1].value == company and row[title_col - 1].value == job_title:
            ws.delete_rows(row[0].row)

    # 在表头下方插入一行新记录，已有记录自动下移（最新公司永远在最上面）
    ws.insert_rows(2)
    r = 2

    values = {
        "职位名称": job_title,
        "公司": company,
        "总体匹配度": overall_match,
        "投递日期": apply_date or date.today(),
        "职位内容": _bullets_plain(job_content_bullets),
        "任职要求": _bullets_rich(requirement_items),
        "技能匹配度-匹配的": _bullets_plain(skill_matched_bullets),
        "技能匹配度-未达标的": _bullets_plain(skill_gap_bullets),
        "相关经验年限": experience_years,
        "行业背景": _bullets_plain(industry_bullets),
        "薪资范围": salary,
        "团队规模/汇报线": _bullets_plain(team_bullets),
        "地理位置/远程要求": location,
        "状态/下一步": status,
        "简历优化内容": _bullets_plain(resume_optimization_bullets) if resume_optimization_bullets else "未生成定制简历",
        "简历存储路径": resume_path or "—",
        "Cover Letter": cover_letter or "未生成",
        "公司简介": company_overview or "未获取到公司简介",
    }

    max_lines = 1
    for col_idx, h in enumerate(HEADERS, start=1):
        if h == "职位名称":
            cell = ws.cell(row=r, column=col_idx, value=job_title)
            if job_url:
                cell.hyperlink = job_url
                cell.font = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")
            else:
                cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = BORDER
            continue

        val = values[h]
        cell = ws.cell(row=r, column=col_idx, value=val)
        cell.border = BORDER

        if h == "任职要求":
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            max_lines = max(max_lines, len(requirement_items))
            continue

        cell.font = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        if h == "总体匹配度":
            cell.number_format = "0%"
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif h == "投递日期":
            cell.number_format = "yyyy-mm-dd"
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif isinstance(val, str) and "\n" in val:
            max_lines = max(max_lines, val.count("\n") + 1)

    ws.row_dimensions[r].height = min(500, max(60, 22 * max_lines))

    wb.save(path)
    return path


def update_entry_fields(path, company, job_title, resume_optimization_bullets=None, resume_path=None, cover_letter=None, status=None):
    """按 公司+职位名称 找到追踪表里那一行，只改材料相关的几个单元格，其余原样不动。

    材料生成从AI匹配分析里拆出来之后需要这个函数：用户点"生成定制简历+Cover Letter"时，
    这条职位的分析行早就写好了，只是那几列还是"未生成"。走 add_entry() 重写整行的话，
    得把职位内容/任职要求/技能匹配等等十几个分析字段全部再传一遍——那些数据现在只存在
    追踪表自己那一行里（分析结论没有完整落库），等于要先读回来再原样写回去，中间任何
    一处反解（富文本的红色标记尤其）不完美都会让已有内容悄悄退化。只动四个格子最安全。

    找不到对应行时返回 False（不报错）：追踪表可能被用户挪走或手工删过行，材料本身已经
    落库了（models.update_job_materials），追踪表没同步上不该让整个生成流程失败。
    """
    if not os.path.exists(path):
        return False
    wb = load_workbook(path, rich_text=True)
    ws = wb["JD匹配追踪表"] if "JD匹配追踪表" in wb.sheetnames else wb.active
    _migrate_headers(ws)
    company_col = HEADERS.index("公司") + 1
    title_col = HEADERS.index("职位名称") + 1

    values = {
        "简历优化内容": _bullets_plain(resume_optimization_bullets) if resume_optimization_bullets else "未生成定制简历",
        "简历存储路径": resume_path or "—",
        "Cover Letter": cover_letter or "未生成",
    }
    if status is not None:
        values["状态/下一步"] = status

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[company_col - 1].value != company or row[title_col - 1].value != job_title:
            continue
        max_lines = 1
        for header, val in values.items():
            cell = ws.cell(row=row[0].row, column=HEADERS.index(header) + 1, value=val)
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if isinstance(val, str) and "\n" in val:
                max_lines = max(max_lines, val.count("\n") + 1)
        # 行高只增不减：这一行其它列（任职要求等）的内容还在，按新写入的几列算出来的高度
        # 可能比原来矮，缩回去会把旁边的长文本压得看不全。
        current = ws.row_dimensions[row[0].row].height or 60
        ws.row_dimensions[row[0].row].height = min(500, max(current, 22 * max_lines))
        wb.save(path)
        return True
    return False


def _plain_bullets_to_list(val):
    """反解 _bullets_plain() 写入的 "• xxx\\n• yyy" 格式为 ["xxx", "yyy"]。
    未生成/占位值（"未生成定制简历"、"未生成"、"—"等不带项目符号的值）原样按单条返回。"""
    if not val or not isinstance(val, str):
        return []
    lines = [l.strip() for l in val.split("\n") if l.strip()]
    return [l[2:] if l.startswith("• ") else l for l in lines]


def _rich_requirements_to_items(val):
    """反解 _bullets_rich() 写入的 CellRichText，还原为 [{"text":..., "is_gap":bool}, ...]。"""
    if val is None:
        return []
    if isinstance(val, str):
        return [{"text": l, "is_gap": False} for l in _plain_bullets_to_list(val)]
    items = []
    for block in val:
        text = block.text if hasattr(block, "text") else str(block)
        is_gap = bool(getattr(block, "font", None) and block.font.color and block.font.color.rgb == "FFFF0000")
        text = text.strip()
        if text.startswith("• "):
            text = text[2:]
        if text:
            items.append({"text": text, "is_gap": is_gap})
    return items


def list_entries(path):
    """按追踪表当前的行顺序（最新的在最上面）返回结构化的记录列表，供网页展示用。
    文件不存在时返回空列表。"""
    if not os.path.exists(path):
        return []
    wb = load_workbook(path, rich_text=True)
    ws = wb["JD匹配追踪表"] if "JD匹配追踪表" in wb.sheetnames else wb.active
    idx = {h: i for i, h in enumerate(HEADERS)}

    def _cell(row, header):
        """按表头名安全取值：旧追踪表文件可能还没有后来新加的列（比如"公司简介"），
        行的实际长度可能比当前 HEADERS 短，直接按下标取会越界报错。"""
        i = idx[header]
        return row[i].value if i < len(row) else None

    entries = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        title_cell = row[idx["职位名称"]]
        company = _cell(row, "公司")
        if not company and not title_cell.value:
            continue

        apply_date = row[idx["投递日期"]].value
        if hasattr(apply_date, "date"):  # datetime -> date（去掉openpyxl读回来时带的00:00:00时间部分）
            apply_date = apply_date.date()
        resume_path = row[idx["简历存储路径"]].value
        cover_letter = row[idx["Cover Letter"]].value
        resume_bullets_raw = row[idx["简历优化内容"]].value
        company_overview = _cell(row, "公司简介")

        entries.append(
            {
                "job_title": title_cell.value,
                "job_url": title_cell.hyperlink.target if title_cell.hyperlink else None,
                "company": company,
                "overall_match": row[idx["总体匹配度"]].value,
                "apply_date": apply_date.isoformat() if hasattr(apply_date, "isoformat") else apply_date,
                "job_content_bullets": _plain_bullets_to_list(row[idx["职位内容"]].value),
                "requirement_items": _rich_requirements_to_items(row[idx["任职要求"]].value),
                "skill_matched_bullets": _plain_bullets_to_list(row[idx["技能匹配度-匹配的"]].value),
                "skill_gap_bullets": _plain_bullets_to_list(row[idx["技能匹配度-未达标的"]].value),
                "experience_years": row[idx["相关经验年限"]].value,
                "industry_bullets": _plain_bullets_to_list(row[idx["行业背景"]].value),
                "salary": row[idx["薪资范围"]].value,
                "team_bullets": _plain_bullets_to_list(row[idx["团队规模/汇报线"]].value),
                "location": row[idx["地理位置/远程要求"]].value,
                "status": row[idx["状态/下一步"]].value,
                "resume_optimization_bullets": [] if resume_bullets_raw == "未生成定制简历" else _plain_bullets_to_list(resume_bullets_raw),
                "resume_path": None if resume_path in (None, "—") else resume_path,
                "cover_letter": None if cover_letter in (None, "未生成") else cover_letter,
                "company_overview": None if company_overview in (None, "未获取到公司简介") else company_overview,
            }
        )
    return entries


def list_existing_jobs(path):
    """返回追踪表里已有的 (公司, 职位名称) 元组列表，用于新搜索到的职位去重判断——
    避免同一个职位被反复分析、重复插入表格。如果文件不存在，返回空列表。"""
    if not os.path.exists(path):
        return []
    wb = load_workbook(path, rich_text=True)
    ws = wb["JD匹配追踪表"] if "JD匹配追踪表" in wb.sheetnames else wb.active
    company_col = HEADERS.index("公司") + 1
    title_col = HEADERS.index("职位名称") + 1
    result = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        company = row[company_col - 1].value
        title = row[title_col - 1].value
        if company and title:
            result.append((str(company).strip(), str(title).strip()))
    return result


def update_match_score(path, company, job_title, new_score, status_note=None):
    """更新已存在记录的总体匹配度（比如评分方法调整后需要重新计算历史记录时用），
    按公司+职位名称精确匹配对应行，只改C列（总体匹配度）和可选的N列（状态/下一步），
    不改动其他列、不新增行、不改变行的位置。"""
    wb = load_workbook(path, rich_text=True)
    ws = wb["JD匹配追踪表"] if "JD匹配追踪表" in wb.sheetnames else wb.active

    company_col = HEADERS.index("公司") + 1
    title_col = HEADERS.index("职位名称") + 1
    match_col = HEADERS.index("总体匹配度") + 1
    status_col = HEADERS.index("状态/下一步") + 1

    updated = False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if (
            row[company_col - 1].value == company
            and row[title_col - 1].value == job_title
        ):
            ws.cell(row=row[0].row, column=match_col, value=new_score).number_format = "0%"
            if status_note:
                ws.cell(row=row[0].row, column=status_col, value=status_note)
            updated = True
            break

    if not updated:
        raise ValueError(f"未找到匹配记录：公司={company}，职位名称={job_title}")

    wb.save(path)
    return path
